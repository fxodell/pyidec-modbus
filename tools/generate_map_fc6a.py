#!/usr/bin/env python3
"""
Dev-only: read IDEC FC6A Modbus address map from Excel and emit fc6a_tagmap.json.
Usage: python tools/generate_map_fc6a.py [path_to.xlsx]
Output: src/pyidec_modbus/data/fc6a_tagmap.json
Requires: openpyxl (pip install openpyxl or use dev extra).
"""

import json
import logging
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# Modbus reference ranges -> (table, 0-based offset)
COIL_LO, COIL_HI = 1, 99_999
DISCRETE_LO, DISCRETE_HI = 100_001, 199_999
INPUT_REG_LO, INPUT_REG_HI = 300_001, 399_999
HOLDING_LO, HOLDING_HI = 400_001, 499_999


def ref_to_table_offset(ref: int) -> tuple[str, int]:
    """Convert Modbus reference to (table_name, 0-based offset)."""
    ref = int(ref)
    if COIL_LO <= ref <= COIL_HI:
        return "coil", ref - 1
    if DISCRETE_LO <= ref <= DISCRETE_HI:
        return "discrete_input", ref - DISCRETE_LO
    if INPUT_REG_LO <= ref <= INPUT_REG_HI:
        return "input_register", ref - INPUT_REG_LO
    if HOLDING_LO <= ref <= HOLDING_HI:
        return "holding_register", ref - HOLDING_LO
    raise ValueError(f"Invalid Modbus reference: {ref}")


def normalize_operand_cell(value: str | None) -> str | None:
    """Uppercase and strip; return None if empty."""
    if value is None:
        return None
    s = str(value).strip().upper()
    return s if s else None


def find_header_row(rows: list[tuple], operand_col_hint: str = "OPERAND") -> int:
    """Return 0-based row index of header containing operand-like column."""
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell and operand_col_hint in str(cell).upper():
                return i
    return 0


def parse_simple_sheet(rows: list[tuple], sheet_name: str) -> list[dict]:
    """Parse sheet with columns: Operand, Modbus Address (Decimal). Returns list of entries."""
    entries: list[dict] = []
    if not rows:
        return entries
    header_idx = find_header_row(rows, "OPERAND")
    if header_idx >= len(rows):
        return entries
    header = [str(c).strip().upper() if c else "" for c in rows[header_idx]]
    # Find column indices
    operand_col = None
    modbus_col = None
    for j, h in enumerate(header):
        if "OPERAND" in h or "FC5A" in h or "MICROSMART" in h:
            operand_col = j
        if "MODBUS" in h and "ADDRESS" in h and "DECIMAL" in h:
            modbus_col = j
        if "MODBUS ADDRESS" in h:
            modbus_col = j
    if operand_col is None:
        for j, h in enumerate(header):
            if h and "OPERAND" in h:
                operand_col = j
                break
    if modbus_col is None:
        for j, h in enumerate(header):
            if "ADDRESS" in h or "MODBUS" in h:
                modbus_col = j
                break
    if operand_col is None:
        operand_col = 0
    if modbus_col is None:
        modbus_col = 1
    # Data rows
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if len(row) <= max(operand_col, modbus_col):
            continue
        op_cell = row[operand_col]
        ref_cell = row[modbus_col]
        op = normalize_operand_cell(op_cell)
        if not op:
            continue
        try:
            ref = int(float(ref_cell)) if ref_cell is not None else None
        except (TypeError, ValueError):
            continue
        if ref is None:
            continue
        try:
            table, offset = ref_to_table_offset(ref)
        except ValueError:
            continue
        if offset < 0:
            continue
        width = 1 if table in ("coil", "discrete_input") else 16
        entries.append({
            "operand": op,
            "table": table,
            "offset": offset,
            "width": width,
            "meta": {"sheet": sheet_name},
        })
    return entries


def parse_timer_counter_sheet(rows: list[tuple], sheet_name: str, prefix: str) -> list[dict]:
    """
    Timer/Counter sheets have 3 groups: Contact (.C), Current Value (.CV), Preset (.PV).
    Each group has operand + Modbus address. Expand into .C, .CV, .PV operands.
    """
    entries: list[dict] = []
    if not rows:
        return entries
    header_idx = find_header_row(rows, "OPERAND")
    if header_idx >= len(rows):
        return entries
    header = [str(c).strip().upper() if c else "" for c in rows[header_idx]]
    # Expect columns like: Operand, Address (Contact), Operand, Address (CV), Operand, Address (PV)
    # Or three blocks of (Operand, Modbus Address)
    suffixes = [(".C", "discrete_input"), (".CV", "input_register"), (".PV", "holding_register")]
    block_size = 2  # operand, address per block
    num_blocks = 3
    start_col = 0
    for block in range(num_blocks):
        col_op = start_col + block * block_size
        col_addr = start_col + block * block_size + 1
        suf, table = suffixes[block]
        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            if len(row) <= col_addr:
                continue
            op_cell = row[col_op]
            ref_cell = row[col_addr]
            op = normalize_operand_cell(op_cell)
            if not op or not op.startswith(prefix):
                continue
            try:
                ref = int(float(ref_cell)) if ref_cell is not None else None
            except (TypeError, ValueError):
                continue
            if ref is None:
                continue
            try:
                _table, offset = ref_to_table_offset(ref)
            except ValueError:
                continue
            if offset < 0:
                continue
            width = 1 if table == "discrete_input" else 16
            entries.append({
                "operand": op + suf,
                "table": table,
                "offset": offset,
                "width": width,
                "meta": {"sheet": sheet_name},
            })
    return entries


def sheet_rows(ws) -> list[tuple]:
    """Return all rows from worksheet as list of tuples (values only)."""
    return [tuple(cell.value for cell in row) for row in ws.iter_rows()]


def main() -> int:
    xlsx_paths = [
        Path(__file__).resolve().parent.parent / "data" / "FC6A_ModbusSlave_AddressMap_Xml.xlsx",
        Path("/mnt/data/FC6A_ModbusSlave_AddressMap_Xml.xlsx"),
    ]
    if len(sys.argv) > 1:
        xlsx_paths = [Path(sys.argv[1])]
    xlsx_path = None
    for p in xlsx_paths:
        if p.exists():
            xlsx_path = p
            break
    if not xlsx_path:
        logger.error("No Excel file found. Tried: %s", xlsx_paths)
        return 1
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl required. Install with: pip install openpyxl")
        return 1
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    sheet_names = wb.sheetnames
    if not sheet_names and hasattr(wb, "worksheets") and wb.worksheets:
        sheet_names = [ws.title for ws in wb.worksheets]
    if not sheet_names:
        logger.error("Workbook has no sheets. File may be corrupted or unsupported.")
        return 1
    logger.info("Sheets: %s", sheet_names)
    all_entries: list[dict] = []
    seen_operands: set[str] = set()
    table_counts: dict[str, int] = {}
    SIMPLE_SHEETS = [
        "Input", "Output", "Internal Relay", "Special Internal Relay",
        "Shift Register", "Data Register", "Special Data Register",
    ]
    for idx, sheet_name in enumerate(sheet_names):
        try:
            ws = wb[sheet_name]
        except (KeyError, TypeError):
            ws = wb.worksheets[idx] if hasattr(wb, "worksheets") and idx < len(wb.worksheets) else None
        if ws is None:
            continue
        rows = sheet_rows(ws)
        if not rows:
            continue
        if "Timer" in sheet_name:
            entries = parse_timer_counter_sheet(rows, sheet_name, "T")
        elif "Counter" in sheet_name:
            entries = parse_timer_counter_sheet(rows, sheet_name, "C")
        else:
            entries = parse_simple_sheet(rows, sheet_name)
        for e in entries:
            op = e["operand"]
            if op in seen_operands:
                logger.error("Duplicate operand: %s (sheet %s)", op, sheet_name)
                return 1
            seen_operands.add(op)
            tbl = e["table"]
            table_counts[tbl] = table_counts.get(tbl, 0) + 1
            all_entries.append(e)
    wb.close()
    out_path = Path(__file__).resolve().parent.parent / "src" / "pyidec_modbus" / "data" / "fc6a_tagmap.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_entries, f, indent=2)
    logger.info("Wrote %d entries to %s", len(all_entries), out_path)
    for tbl, count in sorted(table_counts.items()):
        logger.info("  %s: %d", tbl, count)
    return 0


if __name__ == "__main__":
    sys.exit(main())
